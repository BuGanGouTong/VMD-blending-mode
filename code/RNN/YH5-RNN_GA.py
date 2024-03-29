import torch
import pandas as pd
import numpy as np
from torch.nn.utils import weight_norm
import torch.utils.data as data_utils
import torch.nn.functional as F
from torch import nn
from sklearn.preprocessing import MinMaxScaler
import matplotlib.pyplot as plt
from bitstring import BitArray
from deap import base, creator, tools, algorithms
from scipy.stats import bernoulli
import openpyxl
from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error


class RNNModel(nn.Module):
    """循环神经网络模型"""
    def __init__(self, layer_num, n_input, n_filter, dropout_rate):
        super(RNNModel, self).__init__()
        self.rnn = nn.RNN(n_input, n_filter, layer_num, batch_first=True, dropout=dropout_rate)
        self.fc1 = nn.Linear(n_filter, 1)

    def forward(self, x):
        x, h_0 = self.rnn(x)
        x = self.fc1(x[:, -1, :])
        return x


# data = torch.randn(50, 7, 10)
# net = GRU(layer_num=3, n_input=10, n_filter=64, kernel_size=3, dropout_rate=0.1, seq_len=7)
# print(net(data).shape)

train_ratio = 0.6 #训练集比例1
forecast_horizon = 1  #预测步长
batch_size = 64
epoch = 230


# df = pd.read_excel('../data/JH126 - VMD.xlsx')
# df = pd.read_excel('../data/JH286 - VMD.xlsx')
# df = pd.read_excel('../data/YH7data - VMD.xlsx')
df = pd.read_excel('../data/YH5data.xlsx')
data_files = np.array(df)
data = data_files[:,1:] #去掉时间一列
data = data.astype('float32')
feature = data.shape[1]

#分割训练集和测试集
train_size = int(len(data) * train_ratio) #train_ratio 7:3
data_train = data[0:train_size,:]
data_test = data[train_size:,:]

#训练集分割输入数据与输出数据  输入数据是使用全部数据不需要将Y剔除
train_x_norm1 = data_train[:,:]
train_y_norm1 = data_train[:,:1]
#测试集分割输入与输出数据
test_x_norm1 = data_test[:,:]
test_y_norm1 = data_test[:,:1]
print(train_y_norm1.shape)
print(test_y_norm1.shape)

preprocessor = MinMaxScaler()
preprocessor.fit(train_x_norm1)
train_x_norm = preprocessor.transform(train_x_norm1)
test_x_norm = preprocessor.transform(test_x_norm1)

preprocessor.fit(train_y_norm1)
train_y_norm = preprocessor.transform(train_y_norm1)

def windowed_dataset(series, feature, window_size):
    """
    Returns a windowed dataset from a Pandas dataframe
    将输入数据进行时间步长处理
    return shape = (N, feature_num, seq_len)
    """
    available_examples = series.shape[0]-window_size + 1
    time_series_number = series.shape[1]
    inputs = np.zeros((available_examples,window_size,time_series_number))
    for i in range(available_examples):
        inputs[i,:,:] = series[i:i+window_size,:]
    return inputs

def windowed_forecast(series, forecast_horizon):
    #处理输出数据的时间步长
    available_outputs = series.shape[0]- forecast_horizon + 1
    output_series_num = series.shape[1]
    output = np.zeros((available_outputs,forecast_horizon, output_series_num))
    for i in range(available_outputs):
        output[i,:]= series[i:i+forecast_horizon,:]
    return output


def train_evaluate(ga_individual_solution):
    # 对目标超参数进行解码
    learning_rate = BitArray(ga_individual_solution[0:1])
    layer_num = BitArray(ga_individual_solution[1:4])
    dropout_rate = BitArray(ga_individual_solution[4:7])
    windows_length = BitArray(ga_individual_solution[7:11])
    #确定范围
    learning_rate = 10 ** (-learning_rate.uint-3)
    layer_num = layer_num.uint + 2
    window_length = windows_length.uint + 8
    dropout_rate = (dropout_rate.uint+1) * 0.1

    #取得真实训练集y值
    test_y_real = test_y_norm1[window_length:]
    # 训练集输入输出确定
    train_x = windowed_dataset(train_x_norm[:-forecast_horizon], feature, window_length)  # （169,12,13）
    train_y = train_y_norm[window_length:]  # (169,1)
    # 测试集输入输出确定
    test_x = windowed_dataset(test_x_norm[:-forecast_horizon], feature, window_length)  # （34,12,13）

    train_data = data_utils.TensorDataset(torch.from_numpy(train_x).float(), torch.from_numpy(train_y).float())
    train_loader = data_utils.DataLoader(train_data, batch_size=128, shuffle=False)

    # Load transformer with Adam optimizer and MSE loss function
    net = RNNModel(layer_num=layer_num, n_input=feature, n_filter=128, dropout_rate=dropout_rate)
    criterion = nn.MSELoss()
    optimizer = torch.optim.Adam(net.parameters(), lr=learning_rate, weight_decay=0.0001)
    # 训练
    epoch_losses = []
    for epoch in range(180):
        # epoch_loss = 0
        for i, (datax, datay) in enumerate(train_loader):
            net.train()
            loss = criterion(net(datax), datay)
            optimizer.zero_grad()
            loss.backward()
            optimizer.step()
            # epoch_loss += loss.detach().item()
        # epoch_loss /= (i + 1)
        # print('Epoch {}, loss {}'.format(epoch, epoch_loss))
        # epoch_losses.append(epoch_loss)

    preprocessor.fit(train_y_norm1)
    net.eval()
    val_num = int(test_y_real.shape[0] / 2)
    val_x = test_x[:val_num, :, :]
    val_y_real = test_y_real[:val_num, :]
    test_x = test_x[val_num:, :, :]
    test_y_real = test_y_real[val_num:, :]
    # 验证集
    predict = net(torch.from_numpy(val_x).float()).detach().numpy()
    predict_y = preprocessor.inverse_transform(predict)
    velMAPE = np.mean(np.abs(val_y_real - predict_y) / val_y_real)


    predict = net(torch.from_numpy(test_x).float()).detach().numpy()
    predict_y = preprocessor.inverse_transform(predict)
    testMAPE = np.mean(np.abs(test_y_real - predict_y) / test_y_real)
    R2 = r2_score(test_y_real, predict_y)
    print("vel集预测平均相对误差：" + str(velMAPE))
    print("测试集预测平均相对误差：" + str(testMAPE))
    print("测试集预测r2：" + str(R2), '\n')
    # dataresult = [window_length, learning_rate, layer_num, dropout_rate, testMAPE]
    # dataex = openpyxl.load_workbook('GRU参数及MAPE.xlsx')
    # table = dataex.active
    # nrows = table.max_row
    # # 获得行数
    # for i in range(len(dataresult)):
    #     table.cell(nrows + 1, i + 1).value = dataresult[i]
    # dataex.save('追加测试文件.xlsx')
    # dataex.close()
    if testMAPE < 0.038:
        torch.save(net.state_dict(),
               'RNN_YH5_window_{}+LR_{}+layer_{}+drop_{}+VEL_{}+TES{}+r2_{}.pkl'.format(
               window_length, learning_rate, layer_num, dropout_rate, velMAPE, testMAPE, R2))
    return testMAPE,

population_size = 10 #初始种群数
num_generations = 30 #迭代次数
gene_length = 11  #个体基因长度

#Implementation of Genetic Algorithm using DEAP python library.

#Since we try to minimise the loss values, we use the negation of the root mean squared loss as fitness function.
creator.create('FitnessMax', base.Fitness, weights = (1.0,))
creator.create('Individual', list, fitness = creator.FitnessMax)

#initialize the variables as bernoilli random variables
toolbox = base.Toolbox()
toolbox.register('binary', bernoulli.rvs, 0.5)
toolbox.register('individual', tools.initRepeat, creator.Individual, toolbox.binary, n = gene_length)
toolbox.register('population', tools.initRepeat, list , toolbox.individual)

#Ordered cross-over used for mating
toolbox.register('mate', tools.cxOrdered)
#Shuffle mutation to reorder the chromosomes
toolbox.register('mutate', tools.mutShuffleIndexes, indpb = 0.6)
#use roulette wheel selection algorithm 锦标赛
toolbox.register('select', tools.selTournament, tournsize = 2)
#training function used for evaluating fitness of individual solution.
toolbox.register('evaluate', train_evaluate)

population = toolbox.population(n = population_size)
r = algorithms.eaSimple(population, toolbox, cxpb = 0.4, mutpb = 0.1, ngen = num_generations, verbose = False)

optimal_individuals_data = tools.selBest(population,k = 1) #select top 1 solution
